# -*- coding: utf-8 -*-
#!/usr/bin/env python3
"""
generate_dashboard.py — ลงทุนดี Portfolio Dashboard Generator
Reads portfolio CSVs → generates portfolio_dashboard.html + portfolio_model.xlsx

Usage:  python generate_dashboard.py
Auto-called by: morning-brief workflow, portfolio summary commands
"""

import csv, json, datetime, io, sys
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT        = Path(__file__).parent
KANT_DIR    = ROOT / "sources/portfolio/kant"
METANG_DIR  = ROOT / "sources/portfolio/me-tang"
PRICES_FILE = ROOT / "live_prices.json"
HTML_OUT    = ROOT / "portfolio_dashboard.html"
XLSX_OUT    = ROOT / "portfolio_model.xlsx"

# ── Allocation Targets ────────────────────────────────────────────────────────
KANT_TARGET = {
    "Core Growth": {"pct": 40, "members": ["VOO","SCHD","JEPQ","KKP NDQ100-UH-E","KKP US500-UH-E","ES-GTECH"]},
    "Thematic":    {"pct": 30, "members": ["UCHINA","UOBSJSM","UEMIF-N","DAOL-RARE","X-NUCTECH","KT-INDIA-D"]},
    "Commodities": {"pct": 15, "members": ["KT-PRECIOUS","KT-MINING","K-AGRI","SIVR","CTVA"]},
    "หุ้นตรงไทย":  {"pct": 15, "members": ["SCB","LAND"]},
}

# ── Geographic Groups ─────────────────────────────────────────────────────────
GEO_GROUPS = {
    "US":              ["VOO","SCHD","JEPQ","KKP NDQ100-UH-E","KKP US500-UH-E","ES-GTECH"],
    "China":           ["UCHINA"],
    "India":           ["KT-INDIA-D"],
    "ASEAN/EM":        ["UOBSJSM","UEMIF-N"],
    "Global Thematic": ["DAOL-RARE","X-NUCTECH"],
    "Precious Metals": ["KT-PRECIOUS","SIVR"],
    "Materials":       ["KT-MINING"],
    "Agriculture":     ["K-AGRI","CTVA"],
    "Thailand":        ["SCB","LAND"],
}
GEO_COLORS = ["#3b82f6","#ef4444","#f97316","#a855f7","#6366f1","#eab308","#78716c","#22c55e","#06b6d4"]

# ── Risk Rules ────────────────────────────────────────────────────────────────
SL_RULES = {
    "KT-PRECIOUS":-15,"KT-MINING":-15,"K-AGRI":-15,
    "UCHINA":-20,"UOBSJSM":-20,"UEMIF-N":-20,"DAOL-RARE":-20,
    "X-NUCTECH":-20,"KT-INDIA-D":-20,
    "KKP NDQ100-UH-E":-20,"KKP US500-UH-E":-20,"ES-GTECH":-20,
    "VOO":-20,"SCHD":-20,"JEPQ":-20,"CTVA":-20,"SIVR":-20,
    "SCB":-20,"LAND":-25,
}
SINGLE_COUNTRY_FUNDS = {"UCHINA","KT-INDIA-D","UOBSJSM"}
CONCENTRATION_CAP    = 15.0

# ── Dividend/Income Estimates (annual yield %) ────────────────────────────────
INCOME_YIELDS = {
    "SCHD": 3.5, "JEPQ": 9.0, "LAND": 5.5, "VOO": 1.3,
    "CTVA": 1.8, "SIVR": 0.0, "SCB": 5.5,
}

# ── Stress Test Betas (vs market shock) ──────────────────────────────────────
STRESS_BETA = {
    "VOO":1.0,"SCHD":0.8,"JEPQ":0.9,"CTVA":0.6,"SIVR":-0.2,
    "KKP NDQ100-UH-E":1.1,"KKP US500-UH-E":1.0,"ES-GTECH":1.1,
    "UCHINA":1.2,"UOBSJSM":1.0,"UEMIF-N":1.0,"DAOL-RARE":0.8,"X-NUCTECH":1.1,"KT-INDIA-D":1.1,
    "KT-PRECIOUS":-0.3,"KT-MINING":0.7,"K-AGRI":0.4,
    "SCB":0.7,"LAND":0.6,
}

# ── Watchlist (เพิ่มเติมได้ใน sources/watchlist.md) ──────────────────────────
WATCHLIST = [
    {"ticker":"KKP","market":"SET","thesis":"Value bank P/E 10x, P/BV 1.00x, Dividend yield 6.82%, Beta 0.47","catalyst":"Re-rating เมื่ออสังหาฯ คลี่คลาย","status":"Monitor","entry":"ต่ำกว่า P/BV 1.0x"},
    {"ticker":"HMPRO","market":"SET","thesis":"Construction retail recovery play, P/E 13.82x, Target ฿7.48","catalyst":"2H26 — งบภาครัฐ + SME recovery","status":"Monitor","entry":"ต่ำกว่า ฿6.00"},
]

# ── Me-Tang Plan ──────────────────────────────────────────────────────────────
METANG_PLAN = [
    {"ticker":"KKP", "target_pct":35,"target_thb":7000,"status":"ยังไม่ซื้อ"},
    {"ticker":"HMPRO","target_pct":35,"target_thb":7000,"status":"ยังไม่ซื้อ"},
    {"ticker":"Cash", "target_pct":30,"target_thb":6000,"status":"Reserve"},
]

# ── Loaders ───────────────────────────────────────────────────────────────────
def load_prices():
    if PRICES_FILE.exists():
        d = json.loads(PRICES_FILE.read_text(encoding="utf-8"))
        raw = d.get("prices", d)
        # Normalize: fetch_prices.py stores {"LAND": {"price": 9.4, ...}, "THB=X": {"price": 33.5, ...}}
        # generate_dashboard expects {"LAND": 9.4, "THB=X": 33.5}
        px = {}
        for k, v in raw.items():
            px[k] = v.get("price") if isinstance(v, dict) else v
        fx_raw = px.get("THB=X", 33.0)
        return px, float(fx_raw) if fx_raw else 33.0, d.get("fetched_at",""), raw
    return {}, 33.0, None, {}

def read_holdings(path):
    rows = []
    if not path.exists(): return rows
    with open(path, newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            t = r["Ticker"].strip()
            if t:
                rows.append({"ticker":t,"shares":float(r["Shares"]),
                             "avg":float(r["Avg_Price"]),"ccy":r.get("Currency","USD").strip()})
    return rows

def read_funds(path):
    rows = []
    if not path.exists(): return rows
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        # detect current NAV column: prefer "Current_NAV", else any col starting with "NAV"
        nav_col = None
        for col in (reader.fieldnames or []):
            c = col.strip()
            if c == "Current_NAV":
                nav_col = col; break
            if c.startswith("NAV") and c != "Avg_NAV":
                nav_col = col
        for r in reader:
            n = r["Fund_Name"].strip()
            if n:
                cnav = float(r[nav_col]) if nav_col and r.get(nav_col,"").strip() else None
                rows.append({"name":n,"units":float(r["Units"]),"avg_nav":float(r["Avg_NAV"]),"current_nav":cnav})
    return rows

def read_transactions(path):
    txns = []
    if not path.exists(): return txns
    with open(path, newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            txns.append(r)
    return txns

# ── Compute Positions ─────────────────────────────────────────────────────────
def holding_val(h, prices, fx):
    t = h["ticker"]
    if t == "CASH":
        return h["shares"], h["shares"], 0.0, h["avg"], h["avg"]
    live = prices.get(f"{t}.BK") or prices.get(t) or prices.get(f"{t}.US")
    cur  = float(live) if live else h["avg"]
    mult = fx if h["ccy"] == "USD" else 1.0
    return h["shares"]*cur*mult, h["shares"]*h["avg"]*mult, (cur-h["avg"])/h["avg"]*100 if h["avg"] else 0, h["avg"], cur

def fund_val(f):
    cur  = f["current_nav"] or f["avg_nav"]
    val  = f["units"] * cur
    cost = f["units"] * f["avg_nav"]
    pnl  = (cur - f["avg_nav"]) / f["avg_nav"] * 100 if f["avg_nav"] else 0
    return val, cost, pnl

def build_positions(holdings_path, funds_path, prices, fx):
    positions = []
    for h in read_holdings(holdings_path):
        val, cost, pnl, avg, cur = holding_val(h, prices, fx)
        positions.append({"name":h["ticker"],"val":val,"cost":cost,"pnl":pnl,
                          "avg":avg,"cur":cur,"ccy":h["ccy"],"shares":h["shares"],
                          "type":"cash" if h["ticker"]=="CASH" else "stock",
                          "sl":SL_RULES.get(h["ticker"],-20)})
    for f in read_funds(funds_path):
        val, cost, pnl = fund_val(f)
        positions.append({"name":f["name"],"val":val,"cost":cost,"pnl":pnl,
                          "avg":f["avg_nav"],"cur":f["current_nav"] or f["avg_nav"],"ccy":"THB",
                          "type":"fund","units":f["units"],"sl":SL_RULES.get(f["name"],-20)})
    return positions

# ── Compute Analytics ─────────────────────────────────────────────────────────
def group_allocations(positions):
    total  = sum(p["val"] for p in positions)
    lookup = {p["name"]:p["val"] for p in positions}
    groups = {}
    for g, info in KANT_TARGET.items():
        gval   = sum(lookup.get(t,0) for t in info["members"])
        actual = gval/total*100 if total else 0
        groups[g] = {"target":info["pct"],"actual":round(actual,1),"variance":round(actual-info["pct"],1),"val":gval}
    return groups, total

def geo_exposure(positions):
    total  = sum(p["val"] for p in positions if p["type"] != "cash")
    lookup = {p["name"]:p["val"] for p in positions}
    result = {}
    for geo, members in GEO_GROUPS.items():
        gval = sum(lookup.get(t,0) for t in members)
        result[geo] = {"val":gval,"pct":round(gval/total*100,1) if total else 0}
    return result

def currency_split(positions):
    usd = sum(p["val"] for p in positions if p.get("ccy")=="USD" and p["type"]!="cash")
    thb = sum(p["val"] for p in positions if p.get("ccy")=="THB" and p["type"]!="cash")
    total = usd + thb
    return {"USD":{"val":usd,"pct":round(usd/total*100,1) if total else 0},
            "THB":{"val":thb,"pct":round(thb/total*100,1) if total else 0}}

def income_estimate(positions):
    total_annual = 0
    items = []
    for p in positions:
        yld = INCOME_YIELDS.get(p["name"], 0)
        if yld > 0:
            ann = p["val"] * yld / 100
            total_annual += ann
            items.append({"name":p["name"],"yield":yld,"annual":ann,"monthly":ann/12})
    return items, total_annual

def stress_test(positions):
    scenarios = {"ตลาดร่วง -10%":-10,"ตลาดร่วง -20%":-20,"ตลาดร่วง -30%":-30}
    results = {}
    total_cost = sum(p["cost"] for p in positions if p["type"] in ("stock","fund"))
    for label, shock_pct in scenarios.items():
        loss = 0
        for p in positions:
            if p["type"] in ("stock","fund"):
                beta = STRESS_BETA.get(p["name"], 1.0)
                loss += p["val"] * (shock_pct/100) * beta
        results[label] = {"loss":loss,"new_val":sum(p["val"] for p in positions)+loss,
                          "loss_pct":loss/sum(p["val"] for p in positions)*100 if sum(p["val"] for p in positions) else 0}
    return results

def build_alerts(positions, groups, total):
    alerts = []
    lookup = {p["name"]:p for p in positions}
    for p in positions:
        if p["type"] in ("stock","fund"):
            sl, pnl = p["sl"], p["pnl"]
            if   pnl <= sl:    alerts.append({"level":"red",   "icon":"🔴","label":"KILL CONDITION","msg":f"{p['name']}: {pnl:.1f}% (SL: {sl}%)"})
            elif pnl <= sl+3:  alerts.append({"level":"orange","icon":"🟠","label":"STOP LOSS NEAR","msg":f"{p['name']}: {pnl:.1f}% (SL: {sl}%)"})
    for fund in SINGLE_COUNTRY_FUNDS:
        if fund in lookup and total > 0:
            w = lookup[fund]["val"]/total*100
            if w > CONCENTRATION_CAP:
                alerts.append({"level":"orange","icon":"🟠","label":"CONCENTRATION BREACH","msg":f"{fund}: {w:.1f}% > {CONCENTRATION_CAP}% cap"})
    for g, info in groups.items():
        if abs(info["variance"]) > 10:
            d = "Overweight" if info["variance"] > 0 else "Underweight"
            alerts.append({"level":"yellow","icon":"🟡","label":"ALLOCATION DRIFT","msg":f"{g}: {d} {abs(info['variance']):.1f}%"})
    if not alerts:
        alerts.append({"level":"green","icon":"🟢","label":"ALL CLEAR","msg":"ไม่มีอะไรน่าเป็นห่วง"})
    return alerts

def health_score(alerts):
    s = 100
    s -= sum(1 for a in alerts if a["label"]=="KILL CONDITION")       *20
    s -= sum(1 for a in alerts if a["label"]=="STOP LOSS NEAR")       *8
    s -= sum(1 for a in alerts if a["label"]=="CONCENTRATION BREACH") *10
    s -= sum(1 for a in alerts if a["label"]=="ALLOCATION DRIFT")     *5
    s  = max(0,min(100,s))
    grade = "A" if s>=85 else ("B" if s>=70 else ("C" if s>=55 else "D"))
    color = {"A":"#10b981","B":"#3b82f6","C":"#f59e0b","D":"#ef4444"}[grade]
    return s, grade, color

def action_plan(alerts, positions):
    actions = []
    for a in alerts:
        if   a["label"]=="KILL CONDITION":      actions.append(f"<strong class='text-red-400'>⚡ ทำทันที:</strong> {a['msg']} — ขาย 50% ก่อน ประเมิน thesis ใหม่ใน 1 สัปดาห์")
        elif a["label"]=="STOP LOSS NEAR":      actions.append(f"<strong class='text-orange-400'>🔔 ภายใน 1 สัปดาห์:</strong> {a['msg']} — เตรียม cut loss plan")
        elif a["label"]=="CONCENTRATION BREACH":actions.append(f"<strong class='text-orange-400'>⚠ ภายใน 2 สัปดาห์:</strong> {a['msg']} — Rebalance ลดสัดส่วน ย้ายไปกลุ่ม Underweight")
        elif a["label"]=="ALLOCATION DRIFT":    actions.append(f"<strong class='text-yellow-400'>📊 ภายใน 1 เดือน:</strong> {a['msg']} — Review rebalancing plan")
    trailing = [p for p in positions if p.get("pnl",0)>=30 and p["type"] in ("stock","fund")]
    if trailing:
        actions.append(f"<strong class='text-green-400'>🎯 Trailing Stop:</strong> {', '.join(p['name'] for p in trailing)} มีกำไร ≥ 30% — ตั้ง Floor = peak × 0.90")
    if not any(a["level"] in ("red","orange") for a in alerts):
        actions.insert(0,"<strong class='text-green-400'>✅ พอร์ตสถานะดี:</strong> ไม่มี urgent alert — ทำตาม monthly review routine")
    actions.append("<strong class='text-blue-400'>📅 ทุกเดือน:</strong> รัน <code class='bg-gray-700 px-1 rounded text-xs'>python fetch_prices.py && python generate_dashboard.py</code>")
    actions.append("<strong class='text-blue-400'>📅 ทุกเดือน:</strong> ตรวจ Allocation Variance — กลุ่มไหนเบี่ยง ±10% → Rebalance ภายใน 2 สัปดาห์")
    actions.append("<strong class='text-purple-400'>🔍 Q3 2026:</strong> Me-Tang — execute KKP + HMPRO entry ตาม plan ถ้า technical จังหวะดี")
    return actions

# ── Position Verdict ──────────────────────────────────────────────────────────
REPLACEMENT_MAP = {
    "Core Growth": "VOO หรือ SCHD (US index กองทุนหลัก)",
    "Thematic":    "DAOL-RARE หรือ X-NUCTECH (Thematic ที่ thesis ดีกว่า)",
    "Commodities": "KT-PRECIOUS หรือ KT-MINING (Commodities ที่ยังไม่ hit SL)",
    "หุ้นตรงไทย": "SCB ถ้า thesis ยังดี หรือย้ายไป Core Growth",
}

def position_verdict(p, groups):
    """Return (icon, color, text) — actionable verdict for each position."""
    pnl, sl, name = p["pnl"], p["sl"], p["name"]

    # Which group is most underweight → replacement suggestion
    uw_group = min(groups.items(), key=lambda x: x[1]["variance"])[0]
    replace  = REPLACEMENT_MAP.get(uw_group, f"กลุ่ม {uw_group}")
    dist     = pnl - sl  # positive = buffer remaining before SL

    if pnl <= sl:
        return ("⚡", "text-red-400 font-bold",
                f"ขาย 50% ทันที — ชน Kill Condition ({pnl:.1f}% ≤ {sl}%) แล้วย้ายเงินไปซื้อ {replace}")
    elif pnl <= sl + 3:
        return ("🔔", "text-orange-400 font-semibold",
                f"เฝ้าระวัง — เหลือ buffer {dist:.1f}% ก่อนถึง SL ถ้าหล่นอีกให้ขายออก แล้วย้ายไป {replace}")
    elif pnl >= 30:
        return ("🎯", "text-emerald-400 font-semibold",
                f"ล็อกกำไร — กำไร {pnl:.1f}% แล้ว ตั้ง trailing stop (floor = peak × 0.90) ยังไม่ต้องขาย")
    elif pnl >= 15:
        return ("✅", "text-green-400",
                f"ถือต่อ — กำไร {pnl:.1f}% thesis ยังดี รอ catalyst ต่อ ไม่ต้องทำอะไร")
    elif pnl >= 0:
        return ("✅", "text-green-400",
                f"ถือต่อ — กำไรเล็กน้อย {pnl:.1f}% ยังอยู่ใน zone ปกติ")
    elif dist > 10:
        return ("👁", "text-gray-400",
                f"เฝ้าดู — ขาดทุน {pnl:.1f}% แต่ยังห่าง SL อีก {dist:.1f}% ยังไม่ต้องทำอะไร")
    else:
        return ("⚠", "text-yellow-400 font-semibold",
                f"ระวัง — ขาดทุน {pnl:.1f}% เหลือ buffer {dist:.1f}% เตรียม cut loss plan ถ้า thesis เปลี่ยน")

# ── HTML Helpers ──────────────────────────────────────────────────────────────
def fmt_thb(v):   return f"฿{v:,.0f}"
def fmt_pct(v):   return f"{v:+.1f}%"
def pnl_cls(v):   return "text-green-400" if v >= 0 else "text-red-400"

def badge_var(v):
    if   v >  5: return f'<span class="px-2 py-0.5 rounded text-xs font-bold bg-red-900 text-red-300">▲ OW {v:+.1f}%</span>'
    elif v < -5: return f'<span class="px-2 py-0.5 rounded text-xs font-bold bg-yellow-900 text-yellow-300">▼ UW {v:+.1f}%</span>'
    else:        return f'<span class="px-2 py-0.5 rounded text-xs font-bold bg-green-900 text-green-300">✓ {v:+.1f}%</span>'

def badge_sl(dist):
    if   dist <= 3: return f'<span class="text-xs px-1.5 py-0.5 rounded bg-red-900 text-red-300">⚠ {dist:.1f}%</span>'
    elif dist <= 8: return f'<span class="text-xs px-1.5 py-0.5 rounded bg-yellow-900 text-yellow-300">{dist:.1f}%</span>'
    else:           return f'<span class="text-xs text-gray-500">{dist:.1f}%</span>'

ALERT_BG = {
    "red":   "bg-red-950 border-red-700 text-red-200",
    "orange":"bg-orange-950 border-orange-700 text-orange-200",
    "yellow":"bg-yellow-950 border-yellow-800 text-yellow-200",
    "green": "bg-green-950 border-green-800 text-green-200",
}

# ── HTML Generation ───────────────────────────────────────────────────────────
def generate_html(kant_pos, kant_groups, kant_total, kant_alerts, kant_sc, kant_gr, kant_clr,
                  kant_geo, kant_ccy, kant_income_items, kant_income_total, kant_stress,
                  metang_pos, metang_total, metang_txns,
                  generated_at, has_live, fx, prices_raw=None):

    # ── Computed values ───────────────────────────────────────────────────────
    investable = [p for p in kant_pos if p["type"] in ("stock","fund")]
    sorted_pnl = sorted(investable, key=lambda x: x["pnl"])
    worst3 = sorted_pnl[:3]
    best3  = sorted_pnl[-3:][::-1]
    total_cost = sum(p["cost"] for p in investable)
    total_return_pct = (kant_total - total_cost)/total_cost*100 if total_cost else 0
    near_sl_count = sum(1 for a in kant_alerts if a["label"] in ("KILL CONDITION","STOP LOSS NEAR"))

    # ── Allocation pie data ───────────────────────────────────────────────────
    alloc_labels = list(kant_groups.keys())
    alloc_values = [kant_groups[g]["actual"] for g in alloc_labels]
    alloc_colors = ["#3b82f6","#8b5cf6","#f59e0b","#10b981"]

    # ── Geo pie data ──────────────────────────────────────────────────────────
    geo_labels = [g for g,v in kant_geo.items() if v["pct"] > 0]
    geo_values = [kant_geo[g]["pct"] for g in geo_labels]
    geo_clrs   = GEO_COLORS[:len(geo_labels)]

    # ── Variance table ────────────────────────────────────────────────────────
    var_rows = ""
    for g, info in kant_groups.items():
        var_rows += f"""
        <tr class="border-b border-gray-700 hover:bg-gray-800">
          <td class="py-3 px-4 font-medium text-gray-200">{g}</td>
          <td class="py-3 px-4 text-center text-gray-400">{info['target']}%</td>
          <td class="py-3 px-4 text-center font-bold text-white">{info['actual']}%</td>
          <td class="py-3 px-4">
            <div class="relative h-2 bg-gray-700 rounded-full">
              <div class="absolute h-2 rounded-full bg-blue-500" style="width:{min(info['actual'],100)}%"></div>
              <div class="absolute h-4 w-0.5 bg-white opacity-40 -top-1" style="left:{info['target']}%"></div>
            </div>
          </td>
          <td class="py-3 px-4 text-center">{badge_var(info['variance'])}</td>
          <td class="py-3 px-4 text-right text-gray-300">{fmt_thb(info['val'])}</td>
        </tr>"""

    # ── Fund scorecard ────────────────────────────────────────────────────────
    fund_total_val  = sum(p["val"]  for p in kant_pos if p["type"]=="fund")
    fund_total_cost = sum(p["cost"] for p in kant_pos if p["type"]=="fund")
    fund_total_pnl  = fund_total_val - fund_total_cost
    fund_total_pct  = fund_total_pnl / fund_total_cost * 100 if fund_total_cost else 0
    fund_rows = ""
    for p in sorted([p for p in kant_pos if p["type"]=="fund"], key=lambda x: x["pnl"]):
        dist    = p["pnl"] - p["sl"]
        pnl_thb = p["val"] - p["cost"]
        color_row = 'bg-red-950' if p["pnl"] <= p["sl"] else ('bg-orange-950' if p["pnl"] <= p["sl"]+3 else '')
        vicon, vcls, vtxt = position_verdict(p, kant_groups)
        fund_rows += f"""
        <tr class="border-b border-gray-700 hover:bg-gray-800 text-sm {color_row}">
          <td class="py-2 px-3 font-medium text-white">🏦 {p['name']}</td>
          <td class="py-2 px-3 text-right text-gray-400 font-mono text-xs">{p.get('units',0):.4f}</td>
          <td class="py-2 px-3 text-right text-gray-400 font-mono text-xs">฿{p['avg']:.4f}</td>
          <td class="py-2 px-3 text-right font-mono text-white text-xs">฿{p['cur']:.4f}</td>
          <td class="py-2 px-3 text-right text-gray-300">{fmt_thb(p['val'])}</td>
          <td class="py-2 px-3 text-right font-bold {pnl_cls(p['pnl'])}">{fmt_pct(p['pnl'])}</td>
          <td class="py-2 px-3 text-right {pnl_cls(pnl_thb)} font-semibold">{fmt_thb(pnl_thb)}</td>
          <td class="py-2 px-3 text-right text-gray-500">{p['sl']}%</td>
          <td class="py-2 px-3 text-center">{badge_sl(dist)}</td>
          <td class="py-2 px-3 {vcls} text-xs">{vicon} {vtxt}</td>
        </tr>"""
    fund_rows += f"""
        <tr class="border-t-2 border-gray-500 bg-gray-800 text-sm font-bold">
          <td class="py-2 px-3 text-gray-300 text-xs uppercase tracking-wider">รวมกองทุน</td>
          <td colspan="3"></td>
          <td class="py-2 px-3 text-right text-white">{fmt_thb(fund_total_val)}</td>
          <td class="py-2 px-3 text-right {pnl_cls(fund_total_pct)}">{fmt_pct(fund_total_pct)}</td>
          <td class="py-2 px-3 text-right {pnl_cls(fund_total_pnl)}">{fmt_thb(fund_total_pnl)}</td>
          <td colspan="3"></td>
        </tr>"""

    # ── Stock position rows ───────────────────────────────────────────────────
    stock_total_val  = sum(p["val"]  for p in kant_pos if p["type"]=="stock")
    stock_total_cost = sum(p["cost"] for p in kant_pos if p["type"]=="stock")
    stock_total_pnl  = stock_total_val - stock_total_cost
    stock_total_pct  = stock_total_pnl / stock_total_cost * 100 if stock_total_cost else 0
    pos_rows = ""
    for p in sorted([p for p in kant_pos if p["type"]=="stock"], key=lambda x: x["val"], reverse=True):
        dist    = p["pnl"] - p["sl"]
        pnl_thb = p["val"] - p["cost"]
        ccy_sym = "$" if p["ccy"] == "USD" else "฿"
        vicon, vcls, vtxt = position_verdict(p, kant_groups)
        pos_rows += f"""
        <tr class="border-b border-gray-700 hover:bg-gray-800 text-sm">
          <td class="py-2 px-3 font-mono font-medium text-white">📈 {p['name']}</td>
          <td class="py-2 px-3 text-right text-gray-400 font-mono text-xs">{p.get('shares',0):.4f}</td>
          <td class="py-2 px-3 text-right text-gray-400 font-mono text-xs">{ccy_sym}{p['avg']:.4f}</td>
          <td class="py-2 px-3 text-right font-mono text-white text-xs">{ccy_sym}{p['cur']:.4f}</td>
          <td class="py-2 px-3 text-right font-bold {pnl_cls(p['pnl'])}">{fmt_pct(p['pnl'])}</td>
          <td class="py-2 px-3 text-right {pnl_cls(pnl_thb)} font-semibold">{fmt_thb(pnl_thb)}</td>
          <td class="py-2 px-3 text-right text-gray-500">{p['sl']}%</td>
          <td class="py-2 px-3 text-center">{badge_sl(dist)}</td>
          <td class="py-2 px-3 {vcls} text-xs">{vicon} {vtxt}</td>
        </tr>"""
    pos_rows += f"""
        <tr class="border-t-2 border-gray-500 bg-gray-800 text-sm font-bold">
          <td class="py-2 px-3 text-gray-300 text-xs uppercase tracking-wider">รวมหุ้นตรง</td>
          <td colspan="3"></td>
          <td class="py-2 px-3 text-right {pnl_cls(stock_total_pct)}">{fmt_pct(stock_total_pct)}</td>
          <td class="py-2 px-3 text-right {pnl_cls(stock_total_pnl)}">{fmt_thb(stock_total_pnl)}</td>
          <td class="py-2 px-3 text-right text-gray-400 text-xs" colspan="3">{fmt_thb(stock_total_val)}</td>
        </tr>"""

    # ── Alert cards ───────────────────────────────────────────────────────────
    alert_cards = "\n".join(
        f'<div class="flex items-start gap-3 p-3 rounded-lg border {ALERT_BG.get(a["level"],ALERT_BG["green"])}">'
        f'<span class="text-lg mt-0.5">{a["icon"]}</span>'
        f'<div><div class="font-bold text-xs tracking-widest uppercase">{a["label"]}</div>'
        f'<div class="text-sm mt-0.5">{a["msg"]}</div></div></div>'
        for a in kant_alerts
    )

    # ── Winners / Losers ──────────────────────────────────────────────────────
    def winner_card(p, rank_cls):
        return (f'<div class="bg-gray-800 rounded-lg p-3">'
                f'<div class="font-mono font-bold text-white text-sm">{p["name"]}</div>'
                f'<div class="text-lg font-black {rank_cls}">{fmt_pct(p["pnl"])}</div>'
                f'<div class="text-xs text-gray-500">{fmt_thb(p["val"])}</div></div>')

    best_cards  = "\n".join(winner_card(p,"text-green-400") for p in best3)
    worst_cards = "\n".join(winner_card(p,"text-red-400")   for p in worst3)

    # ── Income rows ───────────────────────────────────────────────────────────
    income_rows = ""
    for it in sorted(kant_income_items, key=lambda x: x["annual"], reverse=True):
        income_rows += f"""
        <tr class="border-b border-gray-700 text-sm">
          <td class="py-2 px-3 font-medium text-white">{it['name']}</td>
          <td class="py-2 px-3 text-center text-yellow-400 font-bold">{it['yield']:.1f}%</td>
          <td class="py-2 px-3 text-right text-gray-300">{fmt_thb(it['annual'])}</td>
          <td class="py-2 px-3 text-right text-gray-400">{fmt_thb(it['monthly'])}</td>
        </tr>"""

    # ── Stress test rows ──────────────────────────────────────────────────────
    stress_rows = ""
    for label, data in kant_stress.items():
        loss_cls = "text-red-400"
        stress_rows += f"""
        <tr class="border-b border-gray-700 text-sm">
          <td class="py-2 px-3 text-gray-300">{label}</td>
          <td class="py-2 px-3 text-right font-bold {loss_cls}">{fmt_thb(data['loss'])}</td>
          <td class="py-2 px-3 text-right {loss_cls}">{data['loss_pct']:.1f}%</td>
          <td class="py-2 px-3 text-right text-white">{fmt_thb(data['new_val'])}</td>
        </tr>"""

    # ── Watchlist cards ───────────────────────────────────────────────────────
    watchlist_cards = ""
    for w in WATCHLIST:
        watchlist_cards += f"""
        <div class="bg-gray-800 rounded-lg p-4">
          <div class="flex items-center justify-between mb-2">
            <div class="font-mono font-bold text-white text-base">{w['ticker']} <span class="text-xs text-gray-500 font-normal">{w['market']}</span></div>
            <span class="text-xs px-2 py-0.5 rounded bg-blue-900 text-blue-300">{w['status']}</span>
          </div>
          <div class="text-xs text-gray-400 mb-1"><span class="text-gray-500">Thesis:</span> {w['thesis']}</div>
          <div class="text-xs text-gray-400 mb-1"><span class="text-gray-500">Catalyst:</span> {w['catalyst']}</div>
          <div class="text-xs text-green-400"><span class="text-gray-500">Entry zone:</span> {w['entry']}</div>
        </div>"""

    # ── Me-Tang plan rows ─────────────────────────────────────────────────────
    current_cash = next((p for p in metang_pos if p["name"]=="CASH"), None)
    current_cash_val = current_cash["val"] if current_cash else 0
    metang_plan_rows = ""
    for item in METANG_PLAN:
        cur_val = current_cash_val if item["ticker"]=="Cash" else 0
        gap = item["target_thb"] - cur_val if item["ticker"]!="Cash" else 0
        status_cls = "text-yellow-400" if item["status"]=="ยังไม่ซื้อ" else "text-green-400"
        metang_plan_rows += f"""
        <tr class="border-b border-gray-700 text-sm">
          <td class="py-2 px-3 font-mono font-bold text-white">{item['ticker']}</td>
          <td class="py-2 px-3 text-center text-gray-300">{item['target_pct']}%</td>
          <td class="py-2 px-3 text-right text-gray-300">{fmt_thb(item['target_thb'])}</td>
          <td class="py-2 px-3 text-right text-gray-500">{fmt_thb(cur_val) if cur_val else '฿0'}</td>
          <td class="py-2 px-3 text-right text-orange-400">{fmt_thb(gap) if gap > 0 else '-'}</td>
          <td class="py-2 px-3 text-center {status_cls}">{item['status']}</td>
        </tr>"""

    # ── Action plan ───────────────────────────────────────────────────────────
    ap_items = action_plan(kant_alerts, kant_pos)
    ap_html  = "\n".join(
        f'<li class="flex gap-3 text-sm"><span class="text-blue-400 font-bold mt-0.5 shrink-0">›</span><span>{item}</span></li>'
        for item in ap_items
    )

    # ── Legend ────────────────────────────────────────────────────────────────
    alloc_legend = "\n".join(
        f'<div class="flex items-center gap-2 text-xs">'
        f'<div class="w-2.5 h-2.5 rounded-full shrink-0" style="background:{alloc_colors[i]}"></div>'
        f'<span class="text-gray-300 flex-1">{g}</span>'
        f'<span class="font-bold text-white">{kant_groups[g]["actual"]}%</span>'
        f'</div>'
        for i, g in enumerate(alloc_labels)
    )
    geo_legend = "\n".join(
        f'<div class="flex items-center gap-2 text-xs">'
        f'<div class="w-2.5 h-2.5 rounded-full shrink-0" style="background:{geo_clrs[i]}"></div>'
        f'<span class="text-gray-300 flex-1">{g}</span>'
        f'<span class="font-bold text-white">{kant_geo[g]["pct"]}%</span>'
        f'</div>'
        for i, g in enumerate(geo_labels)
    )

    # ── Technical Signals (computed before f-string) ──────────────────────────
    def _ma_row(label, ma_val, price_val, ccy_sym):
        if ma_val is None:
            return f'<div class="flex justify-between text-xs"><span class="text-gray-600">{label}</span><span class="text-gray-600">—</span></div>'
        above = price_val > ma_val
        cls   = "text-green-400" if above else "text-red-400"
        arrow = "▲" if above else "▼"
        return f'<div class="flex justify-between text-xs"><span class="text-gray-500">{label}</span><span class="{cls}">{arrow} {ccy_sym}{ma_val:.2f}</span></div>'

    tech_cards = ""
    if prices_raw:
        for p in sorted([x for x in kant_pos if x["type"]=="stock"], key=lambda x: x["name"]):
            t         = p["name"]
            r         = prices_raw.get(f"{t}.BK") or prices_raw.get(t)
            if not r or not isinstance(r, dict): continue
            price_val = r.get("price")
            if price_val is None: continue
            chg       = r.get("change_pct") or 0
            ind       = r.get("indicators") or {}
            rsi       = ind.get("rsi")
            ma20      = ind.get("ma20")
            ma50      = ind.get("ma50")
            ma200     = ind.get("ma200")
            ccy_sym   = "$" if p["ccy"]=="USD" else "฿"
            chg_cls   = "text-green-400" if chg >= 0 else "text-red-400"
            chg_sign  = "+" if chg >= 0 else ""
            rsi_pct   = f"{min(max(rsi,0),100):.0f}" if rsi is not None else "50"
            rsi_bar   = "bg-red-500" if rsi and rsi > 70 else ("bg-green-500" if rsi and rsi < 30 else "bg-yellow-500")
            rsi_txt   = "text-red-400" if rsi and rsi > 70 else ("text-green-400" if rsi and rsi < 30 else "text-yellow-400")
            rsi_lbl   = "Overbought ⚠" if rsi and rsi > 70 else ("Oversold — Opportunity" if rsi and rsi < 30 else "Neutral")
            rsi_str   = f"{rsi:.1f}" if rsi is not None else "N/A"
            above200  = (price_val > ma200) if ma200 is not None else None
            sig_cls   = "text-green-400" if above200 else ("text-red-400" if above200 is False else "text-gray-500")
            sig_txt   = "Long-term Bullish" if above200 else ("Below MA200 — Caution" if above200 is False else "No MA200 data")
            tech_cards += f"""
        <div class="bg-gray-800 rounded-lg p-3 border border-gray-700">
          <div class="flex items-center justify-between mb-2">
            <span class="font-mono font-bold text-white">{t}</span>
            <span class="text-xs font-mono text-gray-300">{ccy_sym}{price_val:.2f} <span class="{chg_cls}">({chg_sign}{chg:.2f}%)</span></span>
          </div>
          <div class="mb-2">
            <div class="flex justify-between text-xs mb-1">
              <span class="text-gray-500">RSI(14)</span>
              <span class="font-bold {rsi_txt}">{rsi_str} — {rsi_lbl}</span>
            </div>
            <div class="bg-gray-700 rounded-full h-1.5 relative">
              <div class="h-1.5 rounded-full {rsi_bar}" style="width:{rsi_pct}%"></div>
              <div class="absolute h-3 w-px bg-red-600 opacity-70 -top-0.5" style="left:70%"></div>
              <div class="absolute h-3 w-px bg-green-600 opacity-70 -top-0.5" style="left:30%"></div>
            </div>
          </div>
          <div class="space-y-0.5">
            {_ma_row("MA20", ma20, price_val, ccy_sym)}
            {_ma_row("MA50", ma50, price_val, ccy_sym)}
            {_ma_row("MA200", ma200, price_val, ccy_sym)}
          </div>
          <div class="mt-2 pt-1.5 border-t border-gray-700">
            <span class="text-xs font-semibold {sig_cls}">{sig_txt}</span>
          </div>
        </div>"""

    tech_section = f"""
  <!-- ── Technical Signals ───────────────────────────────────────────────── -->
  <div class="card p-5">
    <div class="section-title">📊 Technical Signals — หุ้นตรง (RSI + Moving Averages)</div>
    <div class="grid grid-cols-2 md:grid-cols-3 lg:grid-cols-4 gap-3">
      {tech_cards if tech_cards else '<p class="text-gray-500 text-sm col-span-4">ไม่มีข้อมูล — รัน <code class=\"bg-gray-700 px-1 rounded\">python fetch_prices.py</code> ก่อน</p>'}
    </div>
    <div class="mt-3 text-xs text-gray-600">* RSI(14) + MA คำนวณจาก 1-year historical data · เส้นแดง=70 (Overbought) · เส้นเขียว=30 (Oversold)</div>
  </div>
""" if prices_raw else ""

    price_warn = (
        '<div class="mb-5 p-3 bg-yellow-950 border border-yellow-800 rounded-lg text-yellow-200 text-sm flex items-center gap-2">'
        '⚠️ <strong>ใช้ราคาทุน (Avg Price)</strong> — หุ้นตรงอาจไม่ใช่ราคาปัจจุบัน '
        'รัน <code class="bg-yellow-900 px-1 rounded">python fetch_prices.py</code> ก่อนเพื่อความแม่นยำ</div>'
        if not has_live else ""
    )

    kant_stocks_val = sum(p["val"] for p in kant_pos if p["type"]=="stock")
    kant_funds_val  = sum(p["val"] for p in kant_pos if p["type"]=="fund")
    total_ret_cls   = "text-green-400" if total_return_pct >= 0 else "text-red-400"

    return f"""<!DOCTYPE html>
<html lang="th">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>ลงทุนดี — Portfolio Dashboard</title>
<script src="https://cdn.tailwindcss.com"></script>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  body {{ background-color:#0f172a; font-family:'Segoe UI',system-ui,sans-serif; }}
  .card {{ background:#1e293b; border:1px solid #334155; border-radius:0.75rem; }}
  .score-ring {{
    width:110px;height:110px;border-radius:50%;position:relative;
    background:conic-gradient({kant_clr} {kant_sc*3.6}deg,#0f172a 0deg);
    display:flex;align-items:center;justify-content:center;
  }}
  .score-ring::before{{content:'';position:absolute;width:82px;height:82px;border-radius:50%;background:#1e293b;}}
  .section-title{{font-size:0.7rem;font-weight:600;color:#6b7280;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:1rem;}}
</style>
</head>
<body class="text-gray-100 min-h-screen p-4 md:p-6">
<div class="max-w-7xl mx-auto space-y-6">

  <!-- ── Header ──────────────────────────────────────────────────────────── -->
  <div class="flex flex-wrap items-center justify-between gap-4">
    <div>
      <h1 class="text-xl font-bold text-white">ลงทุนดี <span class="text-blue-400">Portfolio Dashboard</span></h1>
      <p class="text-gray-500 text-xs mt-0.5">อัปเดต: {generated_at} · USD/THB: {fx:.2f}</p>
    </div>
    <div class="flex gap-4 text-right">
      <div>
        <div class="text-gray-500 text-xs">Kant — มูลค่ารวม</div>
        <div class="text-xl font-bold text-white">{fmt_thb(kant_total)}</div>
      </div>
      <div>
        <div class="text-gray-500 text-xs">Total Return</div>
        <div class="text-xl font-bold {total_ret_cls}">{fmt_pct(total_return_pct)}</div>
      </div>
    </div>
  </div>

  {price_warn}

  <!-- ── KPI Row ─────────────────────────────────────────────────────────── -->
  <div class="grid grid-cols-2 md:grid-cols-4 gap-4">

    <div class="card p-4 flex items-center gap-4">
      <div class="score-ring">
        <div class="relative z-10 text-center">
          <div class="text-xl font-black" style="color:{kant_clr}">{kant_sc}</div>
          <div class="text-xs text-gray-500">/100</div>
        </div>
      </div>
      <div>
        <div class="text-xs text-gray-500 uppercase tracking-widest">Health</div>
        <div class="text-4xl font-black" style="color:{kant_clr}">{kant_gr}</div>
      </div>
    </div>

    <div class="card p-4">
      <div class="text-xs text-gray-500 mb-1">Kant Portfolio</div>
      <div class="text-xl font-bold text-white">{fmt_thb(kant_total)}</div>
      <div class="mt-2 grid grid-cols-2 gap-1 text-xs">
        <div class="bg-gray-800 rounded p-1.5"><div class="text-gray-500">หุ้นตรง</div><div class="font-semibold">{fmt_thb(kant_stocks_val)}</div></div>
        <div class="bg-gray-800 rounded p-1.5"><div class="text-gray-500">กองทุน</div><div class="font-semibold">{fmt_thb(kant_funds_val)}</div></div>
      </div>
    </div>

    <div class="card p-4">
      <div class="text-xs text-gray-500 mb-1">Me-Tang AI Portfolio</div>
      <div class="text-xl font-bold text-white">{fmt_thb(metang_total)}</div>
      <div class="text-xs text-yellow-400 mt-1">💵 100% CASH — รอ deploy {fmt_thb(sum(i["target_thb"] for i in METANG_PLAN if i["ticker"]!="Cash"))}</div>
      <div class="mt-2 bg-gray-800 rounded p-1.5 text-xs"><div class="text-gray-500">เริ่มต้น</div><div class="font-semibold">฿20,000</div></div>
    </div>

    <div class="card p-4">
      <div class="text-xs text-gray-500 mb-1">Risk Summary</div>
      <div class="space-y-2 mt-1">
        <div class="flex justify-between text-sm"><span class="text-gray-400">Positions near SL</span><span class="font-bold text-red-400">{near_sl_count}</span></div>
        <div class="flex justify-between text-sm"><span class="text-gray-400">Total alerts</span><span class="font-bold text-orange-400">{len(kant_alerts)}</span></div>
        <div class="flex justify-between text-sm"><span class="text-gray-400">USD exposure</span><span class="font-bold text-blue-400">{kant_ccy['USD']['pct']}%</span></div>
        <div class="flex justify-between text-sm"><span class="text-gray-400">Annual income est.</span><span class="font-bold text-yellow-400">{fmt_thb(kant_income_total)}</span></div>
      </div>
    </div>

  </div>

  <!-- ── Winners / Losers ────────────────────────────────────────────────── -->
  <div class="grid grid-cols-1 md:grid-cols-2 gap-4">
    <div class="card p-4">
      <div class="section-title">🏆 Top Performers</div>
      <div class="grid grid-cols-3 gap-2">{best_cards}</div>
    </div>
    <div class="card p-4">
      <div class="section-title">⚠ Worst Performers</div>
      <div class="grid grid-cols-3 gap-2">{worst_cards}</div>
    </div>
  </div>

  <!-- ── Alerts ──────────────────────────────────────────────────────────── -->
  <div class="card p-4">
    <div class="section-title">🚨 Risk &amp; Stress Alerts</div>
    <div class="grid grid-cols-1 md:grid-cols-2 gap-2">
      {alert_cards}
    </div>
  </div>

  <!-- ── Two Pie Charts ──────────────────────────────────────────────────── -->
  <div class="grid grid-cols-1 lg:grid-cols-2 gap-6">

    <div class="card p-5">
      <div class="section-title">Asset Allocation — Kant</div>
      <div class="flex items-center gap-6">
        <div class="shrink-0" style="width:160px;height:160px"><canvas id="allocPie"></canvas></div>
        <div class="flex-1 space-y-2">{alloc_legend}</div>
      </div>
    </div>

    <div class="card p-5">
      <div class="section-title">Geographic Exposure — Kant</div>
      <div class="flex items-center gap-6">
        <div class="shrink-0" style="width:160px;height:160px"><canvas id="geoPie"></canvas></div>
        <div class="flex-1 space-y-2">{geo_legend}</div>
      </div>
    </div>

  </div>

  <!-- ── Variance Table ──────────────────────────────────────────────────── -->
  <div class="card p-5">
    <div class="section-title">Variance Analysis — Current vs Target Allocation</div>
    <div class="overflow-x-auto">
      <table class="w-full text-sm">
        <thead><tr class="border-b border-gray-700 text-gray-500 text-xs uppercase">
          <th class="py-2 px-4 text-left">กลุ่ม</th>
          <th class="py-2 px-4 text-center">Target</th>
          <th class="py-2 px-4 text-center">Actual</th>
          <th class="py-2 px-4 text-left w-48">Bar (│=target)</th>
          <th class="py-2 px-4 text-center">Status</th>
          <th class="py-2 px-4 text-right">มูลค่า</th>
        </tr></thead>
        <tbody>{var_rows}</tbody>
      </table>
    </div>
  </div>

  <!-- ── Fund Scorecard ──────────────────────────────────────────────────── -->
  <div class="card p-5">
    <div class="section-title">Fund Scorecard — กองทุนรวมทั้งหมด (เรียงจากแย่→ดี) พร้อมคำแนะนำ</div>
    <div class="text-xs text-yellow-600 mb-3">⚠ NAV ต้องอัปเดตด้วยตนเอง (yfinance ไม่มีข้อมูลกองทุนไทย) — แก้ไขใน sources/portfolio/kant/mutual_funds.csv</div>
    <div class="overflow-x-auto">
      <table class="w-full text-sm">
        <thead><tr class="border-b border-gray-700 text-gray-500 text-xs uppercase">
          <th class="py-2 px-3 text-left">กองทุน</th>
          <th class="py-2 px-3 text-right">หน่วย</th>
          <th class="py-2 px-3 text-right">Avg NAV</th>
          <th class="py-2 px-3 text-right">NAV ปัจจุบัน</th>
          <th class="py-2 px-3 text-right">มูลค่า</th>
          <th class="py-2 px-3 text-right">P&amp;L%</th>
          <th class="py-2 px-3 text-right">P&amp;L (THB)</th>
          <th class="py-2 px-3 text-right">SL</th>
          <th class="py-2 px-3 text-center">ห่าง SL</th>
          <th class="py-2 px-3 text-left min-w-64">คำแนะนำ</th>
        </tr></thead>
        <tbody>{fund_rows}</tbody>
      </table>
    </div>
  </div>

  <!-- ── Stock Holdings ──────────────────────────────────────────────────── -->
  <div class="card p-5">
    <div class="section-title">Stock Holdings — หุ้นตรงทั้งหมด (Kant) พร้อมคำแนะนำ</div>
    <div class="overflow-x-auto">
      <table class="w-full text-sm">
        <thead><tr class="border-b border-gray-700 text-gray-500 text-xs uppercase">
          <th class="py-2 px-3 text-left">Ticker</th>
          <th class="py-2 px-3 text-right">จำนวน</th>
          <th class="py-2 px-3 text-right">ราคาทุน/หน่วย</th>
          <th class="py-2 px-3 text-right">ราคาปัจจุบัน</th>
          <th class="py-2 px-3 text-right">P&amp;L%</th>
          <th class="py-2 px-3 text-right">P&amp;L (THB)</th>
          <th class="py-2 px-3 text-right">SL</th>
          <th class="py-2 px-3 text-center">ห่าง SL</th>
          <th class="py-2 px-3 text-left min-w-64">คำแนะนำ</th>
        </tr></thead>
        <tbody>{pos_rows}</tbody>
      </table>
    </div>
  </div>

  {tech_section}

  <!-- ── Currency + Stress + Income ─────────────────────────────────────── -->
  <div class="grid grid-cols-1 lg:grid-cols-3 gap-6">

    <!-- Currency Exposure -->
    <div class="card p-5">
      <div class="section-title">Currency Exposure</div>
      <div class="space-y-3">
        <div>
          <div class="flex justify-between text-sm mb-1">
            <span class="text-blue-400 font-bold">🇺🇸 USD</span>
            <span class="text-white font-bold">{kant_ccy['USD']['pct']}%</span>
          </div>
          <div class="bg-gray-700 rounded-full h-3">
            <div class="bg-blue-500 h-3 rounded-full" style="width:{kant_ccy['USD']['pct']}%"></div>
          </div>
          <div class="text-xs text-gray-500 mt-0.5">{fmt_thb(kant_ccy['USD']['val'])}</div>
        </div>
        <div>
          <div class="flex justify-between text-sm mb-1">
            <span class="text-cyan-400 font-bold">🇹🇭 THB</span>
            <span class="text-white font-bold">{kant_ccy['THB']['pct']}%</span>
          </div>
          <div class="bg-gray-700 rounded-full h-3">
            <div class="bg-cyan-500 h-3 rounded-full" style="width:{kant_ccy['THB']['pct']}%"></div>
          </div>
          <div class="text-xs text-gray-500 mt-0.5">{fmt_thb(kant_ccy['THB']['val'])}</div>
        </div>
        <div class="pt-2 border-t border-gray-700 text-xs text-gray-500">
          USD/THB exposure ช่วยกระจายความเสี่ยงค่าเงินบาท — target USD 60-70%
        </div>
      </div>
    </div>

    <!-- Stress Test -->
    <div class="card p-5">
      <div class="section-title">Stress Test Scenarios</div>
      <table class="w-full text-sm">
        <thead><tr class="border-b border-gray-700 text-gray-500 text-xs">
          <th class="py-1 text-left">Scenario</th>
          <th class="py-1 text-right">ขาดทุน</th>
          <th class="py-1 text-right">%</th>
          <th class="py-1 text-right">เหลือ</th>
        </tr></thead>
        <tbody>{stress_rows}</tbody>
      </table>
      <div class="mt-3 text-xs text-gray-500">* คำนวณจาก asset-class beta โดยประมาณ — ใช้เป็น reference เท่านั้น</div>
    </div>

    <!-- Income Estimate -->
    <div class="card p-5">
      <div class="section-title">Income Estimate (Annual)</div>
      <div class="text-2xl font-bold text-yellow-400 mb-3">{fmt_thb(kant_income_total)} <span class="text-sm text-gray-500">/ ปี</span></div>
      <table class="w-full text-sm">
        <thead><tr class="border-b border-gray-700 text-gray-500 text-xs">
          <th class="py-1 text-left">สินทรัพย์</th>
          <th class="py-1 text-center">Yield</th>
          <th class="py-1 text-right">ต่อปี</th>
          <th class="py-1 text-right">ต่อเดือน</th>
        </tr></thead>
        <tbody>{income_rows}</tbody>
      </table>
      <div class="mt-2 text-xs text-gray-500">* ประมาณการจาก historical yield — ไม่ใช่การการันตีผลตอบแทน</div>
    </div>

  </div>

  <!-- ── Watchlist ────────────────────────────────────────────────────────── -->
  <div class="card p-5">
    <div class="section-title">👁 Watchlist — รายการที่กำลังติดตาม</div>
    <div class="grid grid-cols-1 md:grid-cols-2 gap-4">
      {watchlist_cards}
    </div>
  </div>

  <!-- ── Me-Tang Plan ─────────────────────────────────────────────────────── -->
  <div class="card p-5">
    <div class="section-title">🤖 Me-Tang AI Portfolio — Plan vs Current</div>
    <div class="mb-4 p-3 bg-yellow-950 border border-yellow-800 rounded-lg text-yellow-200 text-sm">
      ⏳ <strong>ยังไม่ได้ Deploy</strong> — เงิน ฿20,000 ยังเป็น Cash 100%
      · ต้องซื้อ KKP ฿7,000 + HMPRO ฿7,000 เมื่อ technical จังหวะดี
    </div>
    <div class="overflow-x-auto">
      <table class="w-full text-sm">
        <thead><tr class="border-b border-gray-700 text-gray-500 text-xs uppercase">
          <th class="py-2 px-3 text-left">สินทรัพย์</th>
          <th class="py-2 px-3 text-center">Target %</th>
          <th class="py-2 px-3 text-right">Target (THB)</th>
          <th class="py-2 px-3 text-right">ปัจจุบัน</th>
          <th class="py-2 px-3 text-right">Gap</th>
          <th class="py-2 px-3 text-center">สถานะ</th>
        </tr></thead>
        <tbody>{metang_plan_rows}</tbody>
      </table>
    </div>
  </div>

  <!-- ── 90-Day Action Plan ───────────────────────────────────────────────── -->
  <div class="card p-5">
    <div class="section-title">📋 Proactive Action Plan — 90 วัน</div>
    <ul class="space-y-3">
      {ap_html}
    </ul>
  </div>

  <div class="text-center text-xs text-gray-600 pb-4">
    ลงทุนดี Dashboard · Generated {generated_at} · ข้อมูลนี้ไม่ใช่คำแนะนำทางการเงิน
  </div>

</div>

<script>
// Allocation Pie
new Chart(document.getElementById('allocPie'),{{
  type:'doughnut',
  data:{{
    labels:{json.dumps(alloc_labels,ensure_ascii=False)},
    datasets:[{{data:{json.dumps(alloc_values)},backgroundColor:{json.dumps(alloc_colors)},borderColor:'#1e293b',borderWidth:3}}]
  }},
  options:{{responsive:true,maintainAspectRatio:true,cutout:'62%',plugins:{{legend:{{display:false}}}}}}
}});

// Geo Pie
new Chart(document.getElementById('geoPie'),{{
  type:'doughnut',
  data:{{
    labels:{json.dumps(geo_labels,ensure_ascii=False)},
    datasets:[{{data:{json.dumps(geo_values)},backgroundColor:{json.dumps(geo_clrs)},borderColor:'#1e293b',borderWidth:3}}]
  }},
  options:{{responsive:true,maintainAspectRatio:true,cutout:'62%',plugins:{{legend:{{display:false}}}}}}
}});
</script>
</body>
</html>"""

# ── Excel Generation ──────────────────────────────────────────────────────────
def generate_xlsx(kant_pos, kant_groups, kant_total, kant_alerts, metang_pos, metang_total):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference

    wb = Workbook()

    def hdr(cell, text, bg="1E40AF", fg="FFFFFF"):
        cell.value = text
        cell.font  = Font(bold=True, color=fg, size=10)
        cell.fill  = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def tb():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    # ── Sheet 1: Summary ───────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Summary"
    ws1.column_dimensions["A"].width = 28
    for c in "BCD": ws1.column_dimensions[c].width = 16

    ws1["A1"] = "ลงทุนดี — Portfolio Summary"
    ws1["A1"].font = Font(bold=True, size=14, color="1E40AF")
    ws1["A2"] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws1["A2"].font = Font(color="888888", size=9)

    for col, txt in zip("ABCD", ["Portfolio","Value (THB)","Cost (THB)","Return%"]):
        hdr(ws1[f"{col}4"], txt)

    rows = [
        ("Kant — หุ้นตรง",    sum(p["val"]  for p in kant_pos if p["type"]=="stock"),
                               sum(p["cost"] for p in kant_pos if p["type"]=="stock")),
        ("Kant — กองทุนรวม",  sum(p["val"]  for p in kant_pos if p["type"]=="fund"),
                               sum(p["cost"] for p in kant_pos if p["type"]=="fund")),
        ("Kant — รวม",         kant_total,
                               sum(p["cost"] for p in kant_pos if p["type"] in ("stock","fund"))),
        ("Me-Tang",             metang_total, metang_total),
    ]
    for i, (name, val, cost) in enumerate(rows, start=5):
        ws1[f"A{i}"] = name
        ws1[f"B{i}"] = round(val,2);  ws1[f"B{i}"].number_format = '#,##0.00'
        ws1[f"C{i}"] = round(cost,2); ws1[f"C{i}"].number_format = '#,##0.00'
        ws1[f"D{i}"] = f"=(B{i}-C{i})/C{i}"; ws1[f"D{i}"].number_format = "0.00%"
        for col in "ABCD": ws1[f"{col}{i}"].border = tb()

    # Alerts
    ar = len(rows)+6
    ws1[f"A{ar}"] = "Alerts"; ws1[f"A{ar}"].font = Font(bold=True, color="CC0000")
    for j, a in enumerate(kant_alerts, start=ar+1):
        ws1[f"A{j}"] = f"{a['icon']} {a['label']}: {a['msg']}"
        ws1[f"A{j}"].font = Font(color={"red":"FF0000","orange":"FF6600","yellow":"CC8800","green":"007700"}.get(a["level"],"000000"))

    # ── Sheet 2: Kant Holdings ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Kant Holdings")
    for w, c in zip([14,12,12,12,14,10,12], "ABCDEFG"): ws2.column_dimensions[c].width = w
    for col, txt in zip("ABCDEFG", ["Ticker","Avg Price","Current","Shares","Value (THB)","P&L%","Stop-Loss"]):
        hdr(ws2[f"{col}1"], txt)
    sp = [p for p in kant_pos if p["type"]=="stock"]
    for i, p in enumerate(sp, start=2):
        ws2[f"A{i}"] = p["name"]
        units = p["cost"]/p["avg"] if p["avg"] else 0
        ws2[f"B{i}"] = round(p["avg"],4); ws2[f"B{i}"].number_format = '#,##0.00'
        ws2[f"C{i}"] = round(p["cur"],4); ws2[f"C{i}"].number_format = '#,##0.00'
        ws2[f"D{i}"] = round(units,4);    ws2[f"D{i}"].number_format = '#,##0.0000'
        ws2[f"E{i}"] = round(p["val"],2); ws2[f"E{i}"].number_format = '#,##0.00'
        ws2[f"F{i}"] = f"=(C{i}-B{i})/B{i}"; ws2[f"F{i}"].number_format = "0.00%"
        ws2[f"F{i}"].fill = PatternFill("solid", fgColor="C6EFCE" if p["pnl"]>=0 else "FFCCCC")
        ws2[f"G{i}"] = f"{p['sl']}%"
        for col in "ABCDEFG": ws2[f"{col}{i}"].border = tb()
    n2 = len(sp)+2
    ws2[f"A{n2}"] = "TOTAL"; ws2[f"A{n2}"].font = Font(bold=True)
    ws2[f"E{n2}"] = f"=SUM(E2:E{n2-1})"; ws2[f"E{n2}"].number_format = '#,##0.00'; ws2[f"E{n2}"].font = Font(bold=True)

    # ── Sheet 3: Kant Mutual Funds ─────────────────────────────────────────
    ws3 = wb.create_sheet("Kant Mutual Funds")
    for w, c in zip([22,10,12,14,14,10,12], "ABCDEFG"): ws3.column_dimensions[c].width = w
    for col, txt in zip("ABCDEFG", ["Fund","Units","Avg NAV","Current NAV","Value (THB)","P&L%","Stop-Loss"]):
        hdr(ws3[f"{col}1"], txt)
    fp = [p for p in kant_pos if p["type"]=="fund"]
    for i, p in enumerate(fp, start=2):
        units = p["cost"]/p["avg"] if p["avg"] else 0
        ws3[f"A{i}"] = p["name"]
        ws3[f"B{i}"] = round(units,4);    ws3[f"B{i}"].number_format = '#,##0.0000'
        ws3[f"C{i}"] = round(p["avg"],4); ws3[f"C{i}"].number_format = '#,##0.0000'
        ws3[f"D{i}"] = round(p["cur"],4); ws3[f"D{i}"].number_format = '#,##0.0000'
        ws3[f"E{i}"] = round(p["val"],2); ws3[f"E{i}"].number_format = '#,##0.00'
        ws3[f"F{i}"] = f"=(D{i}-C{i})/C{i}"; ws3[f"F{i}"].number_format = "0.00%"
        ws3[f"F{i}"].fill = PatternFill("solid", fgColor="C6EFCE" if p["pnl"]>=0 else "FFCCCC")
        ws3[f"G{i}"] = f"{p['sl']}%"
        for col in "ABCDEFG": ws3[f"{col}{i}"].border = tb()
    n3 = len(fp)+2
    ws3[f"A{n3}"] = "TOTAL"; ws3[f"A{n3}"].font = Font(bold=True)
    ws3[f"E{n3}"] = f"=SUM(E2:E{n3-1})"; ws3[f"E{n3}"].number_format = '#,##0.00'; ws3[f"E{n3}"].font = Font(bold=True)

    # ── Sheet 4: Allocation Analysis + Chart ──────────────────────────────
    ws4 = wb.create_sheet("Allocation Analysis")
    for w, c in zip([18,12,12,12], "ABCD"): ws4.column_dimensions[c].width = w
    for col, txt in zip("ABCD", ["กลุ่ม","Target%","Actual%","Variance%"]): hdr(ws4[f"{col}1"], txt)
    for i, (g, info) in enumerate(kant_groups.items(), start=2):
        ws4[f"A{i}"] = g
        ws4[f"B{i}"] = info["target"]/100; ws4[f"B{i}"].number_format = "0%"
        ws4[f"C{i}"] = info["actual"]/100; ws4[f"C{i}"].number_format = "0.0%"
        ws4[f"D{i}"] = f"=C{i}-B{i}";     ws4[f"D{i}"].number_format = "0.0%"
        v = info["variance"]
        ws4[f"D{i}"].fill = PatternFill("solid", fgColor="FFCCCC" if v>5 else ("FFFACD" if v<-5 else "C6EFCE"))
        for col in "ABCD": ws4[f"{col}{i}"].border = tb()
    chart = BarChart(); chart.type="col"; chart.grouping="clustered"
    chart.title="Target vs Actual"; chart.width=16; chart.height=10
    chart.add_data(Reference(ws4,min_col=2,max_col=3,min_row=1,max_row=5), titles_from_data=True)
    chart.set_categories(Reference(ws4,min_col=1,min_row=2,max_row=5))
    ws4.add_chart(chart, "F1")

    # ── Sheet 5: Reconciliation ────────────────────────────────────────────
    ws5 = wb.create_sheet("Reconciliation")
    ws5.column_dimensions["A"].width = 32; ws5.column_dimensions["B"].width = 18
    ws5["A1"] = "Reconciliation Check"; ws5["A1"].font = Font(bold=True, size=12, color="1E40AF")
    for col, txt in zip("AB", ["Item","Value (THB)"]): hdr(ws5[f"{col}3"], txt)
    recon = [
        ("Kant Holdings (stocks)",           sum(p["val"] for p in kant_pos if p["type"]=="stock")),
        ("Kant Mutual Funds",                 sum(p["val"] for p in kant_pos if p["type"]=="fund")),
        ("Kant Total",                         kant_total),
        ("Me-Tang Total",                      metang_total),
        ("Grand Total",                        kant_total+metang_total),
    ]
    for i, (label, val) in enumerate(recon, start=4):
        ws5[f"A{i}"] = label; ws5[f"B{i}"] = round(val,2); ws5[f"B{i}"].number_format = '#,##0.00'
        for col in "AB": ws5[f"{col}{i}"].border = tb()
    r = len(recon)+5
    ws5[f"A{r}"] = "Cross-check (Stocks+Funds=Total?)"
    ws5[f"B{r}"] = '=IF(ABS(B4+B5-B6)<1,"OK","MISMATCH")'
    ws5[f"B{r}"].font = Font(bold=True)

    wb.save(XLSX_OUT)
    return True

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("=== Lungtundee Dashboard Generator ===")

    prices, fx, fetched_at, prices_raw = load_prices()
    has_live = bool(fetched_at)
    print(f"Prices: {'LIVE (' + fetched_at + ')' if has_live else 'using avg (no live_prices.json)'} | USD/THB={fx:.2f}")

    # Kant
    kant_pos              = build_positions(KANT_DIR/"holdings.csv", KANT_DIR/"mutual_funds.csv", prices, fx)
    kant_groups, kant_total = group_allocations(kant_pos)
    kant_alerts           = build_alerts(kant_pos, kant_groups, kant_total)
    kant_sc, kant_gr, kant_clr = health_score(kant_alerts)
    kant_geo              = geo_exposure(kant_pos)
    kant_ccy              = currency_split(kant_pos)
    kant_inc_items, kant_inc_total = income_estimate(kant_pos)
    kant_stress           = stress_test(kant_pos)

    print(f"Kant: {fmt_thb(kant_total)} | Score: {kant_sc}/100 Grade {kant_gr} | Alerts: {len(kant_alerts)}")

    # Me-Tang
    metang_pos   = build_positions(METANG_DIR/"holdings.csv", METANG_DIR/"mutual_funds.csv", prices, fx)
    metang_total = sum(p["val"] for p in metang_pos)
    metang_txns  = read_transactions(METANG_DIR/"transactions.csv")
    print(f"Me-Tang: {fmt_thb(metang_total)}")

    generated_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    print("Generating portfolio_dashboard.html ...", end=" ")
    html = generate_html(kant_pos, kant_groups, kant_total, kant_alerts, kant_sc, kant_gr, kant_clr,
                         kant_geo, kant_ccy, kant_inc_items, kant_inc_total, kant_stress,
                         metang_pos, metang_total, metang_txns,
                         generated_at, has_live, fx, prices_raw)
    HTML_OUT.write_text(html, encoding="utf-8")
    print("OK")

    print("Generating portfolio_model.xlsx ...", end=" ")
    ok = generate_xlsx(kant_pos, kant_groups, kant_total, kant_alerts, metang_pos, metang_total)
    print("OK" if ok else "FAILED (openpyxl missing?)")

    print(f"\nDone! Open: {HTML_OUT.name}")

if __name__ == "__main__":
    main()
